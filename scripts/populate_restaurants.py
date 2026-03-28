import json
import sys
from pathlib import Path

import openpyxl

# Add project root to path so we can import src/
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from src.parser import parse_restaurant_data

XLSX_PATH = "sheets/restaurants.xlsx"
RESTAURANT_SHEET = "Restaurants"


def populate_restaurants_details(json_path: str) -> None:
    """Read Apify JSON, parse amenities, write to Details sheet."""

    # 1. Load raw Apify data
    with open(json_path, encoding="utf-8") as f:
        raw_items = json.load(f)
    print(f"Loaded {len(raw_items)} restaurants from {json_path}")

    # 2. Parse each station through the pipeline
    restaurants = []
    for item in raw_items:
        try:
            restaurant = parse_restaurant_data(item)
            restaurants.append(restaurant)
        except Exception as e:
            print(f"  Skipped: {e}")
            continue
    print(f"Parsed {len(restaurants)} restaurants successfully")

    # 3. Open existing XLSX
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[RESTAURANT_SHEET]

    # 4. Clear old data (keep headers in row 1)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # 5. Write restaurant amenities
    # Headers: place_id | name | address | phone | website | rating |
    for i, r in enumerate(restaurants, start=2):
        ws.cell(row=i, column=1, value=r.place_id)
        ws.cell(row=i, column=2, value=r.name)
        ws.cell(row=i, column=3, value=r.address)
        ws.cell(row=i, column=4, value=r.phone)
        ws.cell(row=i, column=5, value=r.website)
        ws.cell(row=i, column=6, value=r.rating)
        ws.cell(row=i, column=7, value=r.total_reviews)
        ws.cell(row=i, column=8, value=r.price_level)
        ws.cell(row=i, column=9, value=r.has_24_hours)
        ws.cell(row=i, column=10, value=r.has_delivery)
        ws.cell(row=i, column=11, value=r.has_dine_in)
        ws.cell(row=i, column=12, value=r.has_takeout)
        ws.cell(row=i, column=13, value=r.has_wifi)
        ws.cell(row=i, column=14, value=r.has_reservation)
        ws.cell(row=i, column=15, value=r.latitude)
        ws.cell(row=i, column=16, value=r.longitude)
        ws.cell(row=i, column=17, value=r.google_maps_url)

    # 6. Save
    wb.save(XLSX_PATH)
    print(f"Wrote {len(restaurants)} rows to {RESTAURANT_SHEET}")

    wb.close()
    print("Done")


if __name__ == "__main__":
    # Run
    json_file = sys.argv[1] if len(sys.argv) > 1 else "sheets/restaurants_sample.json"
    populate_restaurants_details(json_file)
    