"""Formatea la hoja Data del XLSX para screenshot de Fiverr."""

from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT = "../sheets/Google Maps Scraper - local.xlsx"
OUTPUT = "../sheets/fiverr_sample.xlsx"
SHEET = "Data"

wb = load_workbook(INPUT)
ws = wb[SHEET]

# --- Estilos ---
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

data_font = Font(name="Calibri", size=10)
data_align = Alignment(vertical="center")
number_align = Alignment(horizontal="center", vertical="center")

thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
score_high = Font(name="Calibri", size=10, bold=True, color="217346")
score_mid = Font(name="Calibri", size=10, color="BF8F00")
score_low = Font(name="Calibri", size=10, color="C00000")

# --- Headers ---
max_col = 10  # Solo las 10 columnas con datos reales
for col in range(1, max_col + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = thin_border

# Renombrar headers para que se vean profesionales
headers = [
    "Station ID", "Station Name", "Address", "Latitude", "Longitude",
    "Total Reviews", "Open 24h", "Google Maps URL", "24h Verified", "Score"
]
for i, name in enumerate(headers, 1):
    ws.cell(row=1, column=i).value = name

# --- Filas de datos ---
max_row = ws.max_row
for row in range(2, max_row + 1):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.border = thin_border

        if col in (4, 5, 6, 10):
            cell.alignment = number_align
        else:
            cell.alignment = data_align

    # Filas alternas con color
    if row % 2 == 0:
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).fill = alt_fill

    # Color del Score (columna 10)
    score_cell = ws.cell(row=row, column=10)
    if score_cell.value is not None:
        try:
            score = int(score_cell.value)
            if score >= 80:
                score_cell.font = score_high
            elif score >= 50:
                score_cell.font = score_mid
            else:
                score_cell.font = score_low
        except (ValueError, TypeError):
            pass

# --- Ancho de columnas ---
widths = {
    1: 30,   # Station ID
    2: 30,   # Station Name
    3: 55,   # Address
    4: 12,   # Latitude
    5: 12,   # Longitude
    6: 14,   # Total Reviews
    7: 10,   # Open 24h
    8: 45,   # Google Maps URL
    9: 12,   # 24h Verified
    10: 8,   # Score
}
for col, width in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# --- Eliminar columnas vacias (11-21) ---
for col in range(21, 10, -1):
    ws.delete_cols(col)

# Fila de headers mas alta
ws.row_dimensions[1].height = 30

# Congelar panel para que headers siempre se vean
ws.freeze_panes = "A2"

wb.save(OUTPUT)
print(f"Guardado en {OUTPUT}")
print(f"Filas: {max_row - 1} | Columnas: {max_col}")
