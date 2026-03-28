import openpyxl

# constants for XLSX path and sheet
XLSX_PATH = "sheets/restaurants.xlsx"
RESTAURANT_SHEET = "Restaurants"

# create new workbook
wb = openpyxl.Workbook()

# rename sheet to "Restaurants"
ws = wb.active
ws.title = RESTAURANT_SHEET

# add 17 headers to row 1
ws["A1"] = "place_id"
ws["B1"] = "name"
ws["C1"] = "address"
ws["D1"] = "phone"
ws["E1"] = "website"
ws["F1"] = "rating"
ws["G1"] = "total_reviews"
ws["H1"] = "price_level"
ws["I1"] = "has_24_hours"
ws["J1"] = "has_delivery"
ws["K1"] = "has_dine_in"
ws["L1"] = "has_takeout"
ws["M1"] = "has_wifi"
ws["N1"] = "has_reservation"
ws["O1"] = "latitude"
ws["P1"] = "longitude"
ws["Q1"] = "google_maps_url"

# save workbook
wb.save(XLSX_PATH)
print(f"Created {XLSX_PATH}")

# close workbook
wb.close()