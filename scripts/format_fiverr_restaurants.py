"""Format XLSX sheet for Fiverr screenshot."""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT = "sheets/restaurants.xlsx"
OUTPUT = "sheets/fiverr_restaurants.xlsx"
SHEET = "Restaurants"

wb = load_workbook(INPUT)
ws = wb[SHEET]

# Styles
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

data_font = Font(name="Calibri", size=10)
data_align = Alignment(vertical="center")
number_align = Alignment(horizontal="center", vertical="center")

thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Headers style
max_col = 17
for col in range(1, max_col + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = thin_border

# Rename headers for Fiverr
headers = [
    "Place ID", "Name", "Address", "Phone", "Website", "Rating", "Reviews",
    "Level", "24 Hours", "Delivery", "Dine-in", "Takeout",
    "Wifi", "Reservations", "Latitude", "Longitude", "Google Maps URL"
]
for i, name in enumerate(headers, 1):
    ws.cell(row=1, column=i).value = name

# block 1 - Zebra + data styles
alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

# --- data rows ---
max_row = ws.max_row
for row in range(2, max_row + 1):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = data_align

    # even rows = altern color
    if row % 2 == 0:
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).fill = alt_fill

    # block 2 - Color for rating column (column 6 is rating)
    rating_cell = ws.cell(row=row, column=6)
    if rating_cell.value is not None:
        try:
            rating = float(rating_cell.value)
            if rating >= 4:
                rating_cell.font = Font(color="00FF00") # green
            elif rating >= 3:
                rating_cell.font = Font(color="FFFF00") # yellow
            else:
                rating_cell.font = Font(color="FF0000") # red
        except ValueError:
            pass


# block 3 Width of columns
widths = {
    1: 30,   # Place ID
    2: 30,   # Name
    3: 55,   # Address
    4: 15,   # Phone
    5: 45,   # Website
    6: 8,    # Rating
    7: 10,   # Reviews
    8: 10,   # Level
    9: 10,   # 24 Hours
    10: 10,  # Delivery
    11: 10,  # Dine-in
    12: 10,  # Takeout
    13: 10,  # Wifi
    14: 12,  # Reservations
    15: 12,  # Latitude
    16: 12,  # Longitude
    17: 45,  # Google Maps URL
}
for col, width in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# block 4 - Freeze + save
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"
wb.save(OUTPUT)
print(f"Saved to {OUTPUT}")


