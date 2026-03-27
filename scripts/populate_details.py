"""Populate the Details sheet in the XLSX with amenity data from Apify JSON."""

import json
import sys
from pathlib import Path

import openpyxl

# Add project root to path so we can import src/
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from src.parser import parse_apify_data

XLSX_PATH = "sheets/Google Maps Scraper - local.xlsx"
DETAILS_SHEET = "Details"


def populate_details(json_path: str) -> None:
    """Read Apify JSON, parse amenities, write to Details sheet."""

    # 1. Load raw Apify data
    with open(json_path, encoding="utf-8") as f:
        raw_items = json.load(f)
    print(f"Loaded {len(raw_items)} stations from {json_path}")

    # 2. Parse each station through our pipeline
    stations = []
    for item in raw_items:
        try:
            station = parse_apify_data(item)
            stations.append(station)
        except Exception as e:
            print(f"  Skipped: {e}")
            continue
    print(f"Parsed {len(stations)} stations successfully")

    # 3. Open existing XLSX
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[DETAILS_SHEET]

    # 4. Clear old data (keep headers in row 1)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # 5. Write station amenities starting at row 2
    # Headers: Station_ID | Station_Name | Has_24_Hours | Has_ATM |
    #          Has_CarWash | Has_Store | Price_Level | Has_Oxxo | Has_Coffee
    for i, s in enumerate(stations, start=2):
        ws.cell(row=i, column=1, value=s.Station_ID)
        ws.cell(row=i, column=2, value=s.Station_Name)
        ws.cell(row=i, column=3, value=s.Has_24_Hours)
        ws.cell(row=i, column=4, value=s.Nearby_ATM)
        ws.cell(row=i, column=5, value=s.Has_CarWash)
        ws.cell(row=i, column=6, value=s.Has_Store)
        ws.cell(row=i, column=7, value=s.Precio_Litro)
        ws.cell(row=i, column=8, value=s.Nearby_OXXO)
        ws.cell(row=i, column=9, value=s.Nearby_Coffee)

    # 6. Save
    wb.save(XLSX_PATH)
    print(f"Wrote {len(stations)} rows to '{DETAILS_SHEET}' sheet in {XLSX_PATH}")


if __name__ == "__main__":
    # Default to sample, pass full JSON path as argument later
    json_file = sys.argv[1] if len(sys.argv) > 1 else "sheets/apify_sample.json"
    populate_details(json_file)
